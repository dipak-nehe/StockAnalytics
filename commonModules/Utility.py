import datetime
from pathlib import Path
import openpyxl

# Path("/zacks/data/stocks.xlsx")
base_path = Path(__file__).parent
input_data_project_file_path = (base_path / "../data/stocks.xlsx").resolve()
pathString = Path(input_data_project_file_path)
# Output File Name
output_excel_file_path = (base_path / "../Output/output.").resolve()
output_excel_filename = ""

# % Deviation
deviation = -25.00
stock_vgm_score = ["A", "B"]
stock_rating = ["1", "2"]


# Generate unique output file to write
def unique_output_file(file_name):
    return Path(file_name + str(get_time_stamp()) + ".xlsx")


def read_excel(loc):
    # loc = Path(pathString)
    # print("Path is:" + str(loc))
    wb = openpyxl.load_workbook(loc)
    sheet = wb['Sheet1']
    stocks = []
    for cellObj in list(sheet.columns)[0]:
        # print(sheet.cell_value(i, 0))
        if str(cellObj.value) == "Symbol":
            continue
        stocks.append(str(cellObj.value))
        print(str(cellObj.value))
    return stocks


def write_excel(stockRatingDict):
    global output_excel_filename
    output_excel_filename = unique_output_file(output_excel_file_path)

    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    sheet = wb.active

    # Once have the Worksheet object,
    # one can get its name from the
    # title attribute.
    sheet_title = sheet.title
    # print("active sheet title: " + sheet_title)
    sheet.title = "StockRating"

    # print("active sheet title: " + sheet_title)
    # Note: The first row or column integer
    # is 1, not 0. Cell object is created by
    # using sheet object's cell() method.

    c1 = sheet.cell(row=1, column=1)
    c1.value = "STOCK"
    c2 = sheet.cell(row=1, column=2)
    c2.value = "RATING"

    # iterate over the worksheet object and populate it
    i = 2
    j = 2

    for key in stockRatingDict:
        keyCell = sheet['A' + str(i)]
        valueCell = sheet['B' + str(j)]
        keyCell.value = key
        valueCell.value = stockRatingDict[key]
        i += 1
        j += 1

    # Anytime you modify the Workbook object
    # or its sheets and cells, the spreadsheet
    # file will not be saved until you call
    # the save() workbook method.
    wb.save(output_excel_filename)


def get_time_stamp():
    return datetime.datetime.now().timestamp()


def calculate_percent_less_or_more_the_stock_trading(stock_last_price, week_52_high, week_52_low):
    stock_last_price = float(stock_last_price)
    week_52_high = float(week_52_high)
    week_52_low = float(week_52_low)

    # check the amount of +ve or negative variation w.r.t to 52 Week High
    if week_52_high > stock_last_price:
        percent_less_than_all_time_high = -((week_52_high - stock_last_price) / week_52_high)
    elif week_52_high < stock_last_price:
        percent_less_than_all_time_high = (stock_last_price - week_52_high) / week_52_high
    else:
        percent_less_than_all_time_high = 0.0

    return round(percent_less_than_all_time_high, 4) * 100
